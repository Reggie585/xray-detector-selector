import re
import datetime
import openpyxl
import mysql.connector

EXCEL_FILE = "/Users/reggae/Downloads/xray_detector_database_google_sheets_ready_xsight_updated.xlsx"

DB_NAME = "xray_detector_db"
MYSQL_PASSWORD = "your_password"


def clean_name(value, fallback):
    name = str(value if value else fallback).strip().lower()
    name = re.sub(r"[^a-z0-9]+", "_", name)
    name = name.strip("_")

    if not name:
        name = fallback

    if name[0].isdigit():
        name = "c_" + name

    return name[:60]


def unique_names(names):
    seen = {}
    result = []

    for name in names:
        base = name
        count = seen.get(base, 0)
        seen[base] = count + 1

        if count == 0:
            result.append(base)
        else:
            result.append(f"{base}_{count + 1}")

    return result


def convert_value(value):
    if value == "":
        return None

    if isinstance(value, datetime.datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")

    return value


db = mysql.connector.connect(
    host="127.0.0.1",
    port=3306,
    user="root",
    password=MYSQL_PASSWORD
)

cursor = db.cursor()

cursor.execute(f"CREATE DATABASE IF NOT EXISTS {DB_NAME}")
cursor.execute(f"USE {DB_NAME}")

workbook = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    table_name = clean_name(sheet_name, "sheet")

    rows = list(sheet.iter_rows(values_only=True))

    rows = [
        row for row in rows
        if any(cell is not None and cell != "" for cell in row)
    ]

    if not rows:
        continue

    headers = rows[0]
    column_names = [
        clean_name(header, f"col_{i + 1}")
        for i, header in enumerate(headers)
    ]
    column_names = unique_names(column_names)

    data_rows = rows[1:]

    cursor.execute(f"DROP TABLE IF EXISTS `{table_name}`")

    columns_sql = ", ".join(
        f"`{column}` LONGTEXT NULL"
        for column in column_names
    )

    cursor.execute(f"""
        CREATE TABLE `{table_name}` (
            id INT AUTO_INCREMENT PRIMARY KEY,
            {columns_sql}
        )
    """)

    insert_sql = f"""
        INSERT INTO `{table_name}` ({", ".join(f"`{c}`" for c in column_names)})
        VALUES ({", ".join(["%s"] * len(column_names))})
    """

    clean_rows = []

    for row in data_rows:
        row = list(row)

        while len(row) < len(column_names):
            row.append(None)

        row = row[:len(column_names)]
        row = [convert_value(value) for value in row]
        if any(value is not None for value in row):
            clean_rows.append(row)

    if clean_rows:
        cursor.executemany(insert_sql, clean_rows)

    db.commit()

    print(f"Imported {len(clean_rows)} rows into {table_name}")

cursor.close()
db.close()

print("All data imported successfully.")
